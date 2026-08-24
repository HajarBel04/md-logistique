#!/usr/bin/env python3
"""
MD-Logistique — Module A LBD Tracking (Abdelhakim)
Analyse les colis OSP MD3449 et colore selon le statut de livraison.

Statuts :
  VERT   = No Inbound Scan    — colis jamais inbound scanné au dépôt
  JAUNE  = Retour dépôt       — chauffeur a ramené dans son chariot
  ROUGE  = Driver Error       — dans SCANNING J mais pas revenu
  BLEU   = Fermeture          — non livré cause fermeture hebdomadaire (KC)
  GRIS   = Future delivery    — livraison planifiée
"""

import os
import sys
from datetime import datetime
from io import BytesIO
from typing import Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

OSP_CODE = 'MD3449'

# Couleurs Excel (ARGB sans alpha)
BG_VERT    = 'C6EFCE'
BG_JAUNE   = 'FFEB9C'
BG_ORANGE  = 'FCE4D6'  # Retour dépôt (fichier Retours chauffeur)
BG_ROUGE   = 'FFC7CE'
BG_BLEU    = 'BDD7EE'
BG_GRIS    = 'D9D9D9'

FG_VERT    = '006100'
FG_JAUNE   = '9C6500'
FG_ORANGE  = 'C55A11'  # Retour dépôt
FG_ROUGE   = '9C0006'
FG_BLEU    = '1F4E79'
FG_GRIS    = '595959'

STATUS_LABELS = {
    'vert':   'No Inbound Scan',
    'jaune':  'Retour scanning J+1',
    'orange': 'Retour dépôt',
    'rouge':  'Driver Error',
    'bleu':   'Fermeture',
    'gris':   'Future delivery',
}

COLOR_MAP = {
    'vert':   (BG_VERT,   FG_VERT),
    'jaune':  (BG_JAUNE,  FG_JAUNE),
    'orange': (BG_ORANGE, FG_ORANGE),
    'rouge':  (BG_ROUGE,  FG_ROUGE),
    'bleu':   (BG_BLEU,   FG_BLEU),
    'gris':   (BG_GRIS,   FG_GRIS),
}

STATUS_ORDER = {'rouge': 0, 'orange': 1, 'jaune': 2, 'bleu': 3, 'vert': 4, 'gris': 5}


# ─── Loaders ─────────────────────────────────────────────────────────────────

def _sheet_name_for_date(d: datetime) -> str:
    """2026-07-23 → '23jul'"""
    return f"{d.day}{d.strftime('%b').lower()}"


def _load_lbd(path: str, target_date: datetime) -> list:
    """Retourne les packages MD3449 du bon onglet."""
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_name = _sheet_name_for_date(target_date)

    # Recherche insensible à la casse (ex: '7aug' == '7AUG')
    real_name = next(
        (s for s in wb.sheetnames if s.lower() == sheet_name.lower()),
        None
    )
    if real_name is None:
        available = ', '.join(wb.sheetnames)
        wb.close()
        raise ValueError(
            f"Onglet '{sheet_name}' non trouvé dans {os.path.basename(path)}. "
            f"Disponibles : {available}"
        )

    ws = wb[real_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    headers = [str(h).strip() if h is not None else '' for h in rows[0]]
    packages = []
    for row in rows[1:]:
        if not any(row):
            continue
        r = dict(zip(headers, row))
        if str(r.get('OSP', '')).strip() == OSP_CODE:
            packages.append(r)

    return packages


def _load_scanning(path: Optional[str]) -> set:
    """Tracking numbers du fichier SCANNING dépôt (1ère colonne, skip header)."""
    if not path or not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    trackings = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # ligne d'en-tête ('dépôt ', ...)
        val = row[0]
        if val and str(val).strip() not in ('', 'dépôt ', 'Tracking'):
            trackings.add(str(val).strip())
    wb.close()
    return trackings


def _load_kc(path: Optional[str]) -> set:
    """Tracking numbers KC (pas d'en-tête, toutes les lignes sont des trackings)."""
    if not path or not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    trackings = {str(row[0]).strip() for row in ws.iter_rows(values_only=True)
                 if row[0] and str(row[0]).strip()}
    wb.close()
    return trackings


def _load_future(path: Optional[str]) -> set:
    """Tracking numbers Future (colonnes: ScanDate | Tracking | FutureDate)."""
    if not path or not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    trackings = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        if row[1]:
            trackings.add(str(row[1]).strip())
    wb.close()
    return trackings


def _load_retours(path: Optional[str]) -> set:
    """
    Extrait tous les tracking numbers (1Z...) du fichier Retours.
    Le fichier est une grille de chariots (Q02, Q03…) — les trackings
    sont dispersés dans toutes les cellules.
    """
    if not path or not os.path.exists(path):
        return set()
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    trackings = set()
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            val = str(cell).strip() if cell else ''
            if val.startswith('1Z') and len(val) > 10:
                trackings.add(val)
    wb.close()
    return trackings


# ─── Logique métier ───────────────────────────────────────────────────────────

def _assign_status(
    tracking: str, pkg: dict,
    scanning_j: set, scanning_j1: set,
    retours: set, kc: set, future: set,
) -> str:
    """
    GRIS   = dans Future
    BLEU   = dans KC (fermeture hebdomadaire)
    ORANGE = dans Retours chauffeur (retour dépôt physique)
    JAUNE  = tour complet J∩J+1 (retour scanning J+1, pas dans retours chauffeur)
    ROUGE  = dans SCANNING J + absent J+1 + absent Retours
    VERT   = pas dans SCANNING J
    """
    if tracking in future:
        return 'gris'
    if tracking in kc:
        return 'bleu'

    in_ret = tracking in retours
    in_j   = tracking in scanning_j  if scanning_j  else False
    in_j1  = tracking in scanning_j1 if scanning_j1 else False

    # Retour physique au dépôt (fichier Retours chauffeur) → orange
    if in_ret:
        return 'orange'

    if scanning_j:
        if not in_j:
            return 'vert'
        # Tour complet : scanné J puis rescanned J+1 → jaune
        if in_j1:
            return 'jaune'
        return 'rouge'

    # Sans SCANNING J : J+1 seul → jaune
    if in_j1:
        return 'jaune'
    return 'vert'


# ─── Génération Excel ─────────────────────────────────────────────────────────

def _fill(bg: str) -> PatternFill:
    return PatternFill(fill_type='solid', fgColor=bg)


def _font(fg: str, bold: bool = False) -> Font:
    return Font(color=fg, bold=bold)


def build_excel(
    packages: list,
    scanning_j: set,
    scanning_j1: set,
    retours: set,
    kc: set,
    future: set,
    target_date: datetime,
) -> Tuple[bytes, dict, list]:
    """Construit le fichier Excel colorisé. Retourne (bytes, summary)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"LBD {target_date.strftime('%d-%m-%Y')}"

    summary = {'total': 0, 'vert': 0, 'jaune': 0, 'orange': 0, 'rouge': 0, 'bleu': 0, 'gris': 0}
    rows_data = []
    seen_trackings = set()  # dédoublonnage

    for pkg in packages:
        tracking = str(pkg.get('Trackingnumber', '')).strip()
        if not tracking or tracking in seen_trackings:
            continue
        seen_trackings.add(tracking)
        status = _assign_status(tracking, pkg, scanning_j, scanning_j1, retours, kc, future)
        summary[status] += 1
        summary['total'] += 1

        street  = str(pkg.get('Street',   '') or '').strip()
        city    = str(pkg.get('City',     '') or '').strip()
        zip_    = str(pkg.get('Zip',      '') or '').strip()
        adresse = ', '.join(p for p in [street, f"{zip_} {city}".strip()] if p)

        scandate = pkg.get('Scandate') or pkg.get('scandate') or ''

        rows_data.append({
            'tracking': tracking,
            'osp':      str(pkg.get('OSP',      '') or '').strip(),
            'scandate': scandate,
            'origin':   str(pkg.get('Origin',   '') or '').strip(),
            'zip':      str(pkg.get('Zip',      '') or '').strip(),
            'city':     city,
            'street':   street,
            'client':   str(pkg.get('Customer', '') or '').strip(),
            'shipper':  str(pkg.get('Shipper',  '') or '').strip(),
            'date':     pkg.get('Date') or pkg.get('date') or '',
            'adresse':  adresse,
            'status':   status,
            'comment':  str(pkg.get('comment',  '') or '').strip(),
        })

    # ── Bloc résumé ────────────────────────────────────────────────────────────
    ws.append(['', f"RAPPORT LBD MD3449 — {target_date.strftime('%d/%m/%Y')}", '', '', '', ''])
    ws.cell(ws.max_row, 2).font = Font(bold=True, size=13)

    ws.append([])

    summary_rows = [
        ('Total',  '',                           summary['total'],   None),
        ('VERT',   STATUS_LABELS['vert'],         summary['vert'],    (BG_VERT,   FG_VERT)),
        ('JAUNE',  STATUS_LABELS['jaune'],        summary['jaune'],   (BG_JAUNE,  FG_JAUNE)),
        ('ORANGE', STATUS_LABELS['orange'],       summary['orange'],  (BG_ORANGE, FG_ORANGE)),
        ('ROUGE',  STATUS_LABELS['rouge'],        summary['rouge'],   (BG_ROUGE,  FG_ROUGE)),
        ('BLEU',   STATUS_LABELS['bleu'],         summary['bleu'],    (BG_BLEU,   FG_BLEU)),
        ('GRIS',   STATUS_LABELS['gris'],         summary['gris'],    (BG_GRIS,   FG_GRIS)),
    ]
    for label, desc, count, colors in summary_rows:
        ws.append(['', label, desc, count, '', ''])
        row_idx = ws.max_row
        if colors:
            bg, fg = colors
            ws.cell(row_idx, 2).fill = _fill(bg)
            ws.cell(row_idx, 2).font = _font(fg, bold=True)
            ws.cell(row_idx, 3).fill = _fill(bg)
            ws.cell(row_idx, 3).font = _font(fg)
        else:
            ws.cell(row_idx, 2).font = Font(bold=True)

    ws.append([])

    # ── En-tête colonnes ───────────────────────────────────────────────────────
    col_headers = ['Trackingnumber', 'OSP', 'Scandate', 'Origin', 'Zip', 'City', 'Street', 'Customer', 'Shipper', 'Date', 'Statut', 'Commentaire']
    ws.append(col_headers)
    hdr_row = ws.max_row
    for col, _ in enumerate(col_headers, 1):
        cell = ws.cell(hdr_row, col)
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill(fill_type='solid', fgColor='2F4F4F')
        cell.alignment = Alignment(horizontal='center')

    for r in rows_data:
        bg, fg = COLOR_MAP[r['status']]
        ws.append([
            r['tracking'],
            r['osp'],
            r['scandate'],
            r['origin'],
            r['zip'],
            r['city'],
            r['street'],
            r['client'],
            r['shipper'],
            r['date'],
            STATUS_LABELS[r['status']],
            r['comment'],
        ])
        data_row = ws.max_row
        for col in range(1, 13):
            ws.cell(data_row, col).fill = _fill(bg)
            ws.cell(data_row, col).font = _font(fg)

    # ── Largeurs colonnes ──────────────────────────────────────────────────────
    for i, w in enumerate([26, 8, 22, 8, 6, 16, 32, 30, 24, 12, 18, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary, rows_data


# ─── Point d'entrée principal ─────────────────────────────────────────────────

def process_lbd(
    lbd_path: str,
    scanning_j_paths: list,   # SCANNING jour J (livraison)
    scanning_j1_paths: list,  # SCANNING jour J+1 (retours tour complet)
    retour_paths: list,       # Retours chariot chauffeur
    kc_paths: list,           # KC fermetures hebdomadaires
    future_path: Optional[str],
    target_date_str: str,
    output_path: Optional[str] = None,
    # legacy
    scanning_paths: list = None,
    kc_j_path: Optional[str] = None,
    kc_retour_paths: list = None,
) -> dict:
    target_date = datetime.strptime(target_date_str, '%Y-%m-%d')

    packages = _load_lbd(lbd_path, target_date)

    scanning_j: set = set()
    for path in (scanning_j_paths or scanning_paths or []):
        scanning_j |= _load_scanning(path)

    scanning_j1: set = set()
    for path in (scanning_j1_paths or []):
        scanning_j1 |= _load_scanning(path)

    retours: set = set()
    for path in (retour_paths or []):
        retours |= _load_retours(path)

    kc: set = set()
    for path in (kc_paths or []):
        kc |= _load_kc(path)

    future = _load_future(future_path)

    excel_bytes, summary, rows = build_excel(
        packages, scanning_j, scanning_j1, retours, kc, future, target_date
    )

    if output_path:
        out_dir = os.path.dirname(output_path)
        if out_dir:
            os.makedirs(out_dir, exist_ok=True)
        with open(output_path, 'wb') as f:
            f.write(excel_bytes)

    return {'summary': summary, 'excel_bytes': excel_bytes, 'rows': rows}


# ─── Filtre MD3449 (préserve dropdowns) ──────────────────────────────────────

COMMENT_VALUES = '"DRIVER ERROR,NO INBOUND SCAN,BABY LABEL,ARRIVED DAY AFTER,OTHER,NO ANSWER"'


def filter_lbd_md3449(lbd_path: str) -> bytes:
    """
    Copie tous les onglets du LBD en ne gardant que les lignes OSP=MD3449.
    Réapplique le dropdown sur la colonne 'comment'.
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    from copy import copy

    src_wb = load_workbook(lbd_path)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    for sheet_name in src_wb.sheetnames:
        src_ws = src_wb[sheet_name]
        src_rows = list(src_ws.iter_rows(values_only=False))
        if not src_rows:
            continue

        header_row = src_rows[0]
        headers = [str(cell.value).strip() if cell.value is not None else '' for cell in header_row]

        if 'OSP' not in headers:
            continue

        osp_idx     = headers.index('OSP')
        comment_idx = headers.index('comment') if 'comment' in headers else None
        md_rows     = [row for row in src_rows[1:] if str(row[osp_idx].value or '').strip() == OSP_CODE]

        if not md_rows:
            continue

        out_ws = out_wb.create_sheet(sheet_name)

        def _copy_row(src_row, dst_ws, dst_row_idx):
            for col_idx, src_cell in enumerate(src_row, 1):
                dst_cell = dst_ws.cell(row=dst_row_idx, column=col_idx, value=src_cell.value)
                if src_cell.has_style:
                    dst_cell.font          = copy(src_cell.font)
                    dst_cell.fill          = copy(src_cell.fill)
                    dst_cell.border        = copy(src_cell.border)
                    dst_cell.alignment     = copy(src_cell.alignment)
                    dst_cell.number_format = src_cell.number_format

        _copy_row(header_row, out_ws, 1)
        for i, row in enumerate(md_rows, 2):
            _copy_row(row, out_ws, i)

        # Dropdown sur la colonne comment
        if comment_idx is not None:
            col_letter = get_column_letter(comment_idx + 1)
            last_row   = 1 + len(md_rows)
            dv = DataValidation(
                type='list',
                formula1=COMMENT_VALUES,
                allow_blank=True,
                showDropDown=False,
            )
            dv.sqref = f'{col_letter}2:{col_letter}{last_row}'
            out_ws.add_data_validation(dv)

        # Largeurs de colonnes
        for col_letter, col_dim in src_ws.column_dimensions.items():
            out_ws.column_dimensions[col_letter].width = col_dim.width

    src_wb.close()

    buf = BytesIO()
    out_wb.save(buf)
    return buf.getvalue()


# ─── CLI ──────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    HAKIM   = os.path.join(BASE, 'samples', 'hakim')
    OUTPUTS = os.path.join(BASE, 'outputs')
    os.makedirs(OUTPUTS, exist_ok=True)

    target = sys.argv[1] if len(sys.argv) > 1 else '2026-07-23'

    result = process_lbd(
        lbd_path          = os.path.join(HAKIM, 'LBDJUL (8) (1).xlsx'),
        scanning_j_paths  = [],
        scanning_j1_paths = [os.path.join(HAKIM, 'SCANNING 280726.xlsx')],
        retour_paths      = [os.path.join(HAKIM, 'Retours 270726.xlsx')],
        kc_paths          = [os.path.join(HAKIM, 'KC 270726.xlsx'),
                             os.path.join(HAKIM, 'KC 280726.xlsx')],
        future_path       = os.path.join(HAKIM, 'Future summer 2026.xlsx'),
        target_date_str   = target,
        output_path       = os.path.join(OUTPUTS, f'LBD_rapport_{target}.xlsx'),
    )

    s = result['summary']
    print(f"\n{'='*52}")
    print(f"  RÉSUMÉ LBD MD3449 — {target}")
    print(f"{'='*52}")
    print(f"  Total  : {s['total']}")
    print(f"  VERT   : {s['vert']:>3}  — {STATUS_LABELS['vert']}")
    print(f"  JAUNE  : {s['jaune']:>3}  — {STATUS_LABELS['jaune']}")
    print(f"  ORANGE : {s['orange']:>3}  — {STATUS_LABELS['orange']}")
    print(f"  ROUGE  : {s['rouge']:>3}  — {STATUS_LABELS['rouge']}")
    print(f"  BLEU   : {s['bleu']:>3}  — {STATUS_LABELS['bleu']}")
    print(f"  GRIS   : {s['gris']:>3}  — {STATUS_LABELS['gris']}")
    print(f"{'='*52}")
    print(f"  Rapport → outputs/LBD_rapport_{target}.xlsx")
