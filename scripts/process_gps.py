#!/usr/bin/env python3
"""
MD-Logistique — Module GPS Analysis
Analyse le fichier GPS du chauffeur :
  - Détection retards express (col L > commit time col G)
  - Distance GPS (col Q/R) vs adresse réelle (col H/I/J/K)
  - Liens Google Maps (Street View + Itinéraire)
  - Règle CLOSED 1/2/3 → commit 17h
  - Règle CP 2040 → commit 14h
"""

import os
import json
import time
import math
from datetime import datetime, time as dtime
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
GEOCACHE    = os.path.join(BASE_DIR, 'outputs', 'geocache.json')

# ─── Couleurs ────────────────────────────────────────────────────────────────

BG_ORANGE      = 'FF8C00'; FG_ORANGE = 'FFFFFF'   # fond orange vif
BG_LATE_EXPRESS = 'FF8C00'; FG_LATE_EXPRESS = 'FF0000'  # fond orange + texte rouge gras
BG_GREEN       = 'C6EFCE'; FG_GREEN  = '006100'
BG_GREEN_LIGHT = 'E2EFDA'; FG_GREEN_LIGHT = '375623'
BG_YELLOW      = 'FFEB9C'; FG_YELLOW = '9C6500'
BG_RED         = 'FFC7CE'; FG_RED    = '9C0006'
BG_GREY        = 'D9D9D9'; FG_GREY   = '595959'

# ─── Codes service UPS (positions 8-9 dans barcode 1Z) ───────────────────────
# Source : UPS Service Level Indicators table

# UPS EXPRESS PLUS → deadline 14h
EXPRESS_PLUS_CODES = {
    '54', '73', 'G2', 'G6', 'G1', 'G5', 'AK', 'V5', 'AM', 'AL', 'HP', 'N5', '5N', '5P',
    '34', 'G3', 'G7', 'G4', 'G8', 'N4', 'P3',
}

# UPS EXPRESS → deadline 14h
EXPRESS_CODES = {
    '66', '75', 'C6', 'C7', 'D3', 'D4', '85', 'V4', '96', '92', 'AS', 'CQ', '5T', '5W',
    '69', '76', 'C9', 'CA', 'D6', 'D7', 'Y6', 'Y9', 'Y7', 'Y8', 'AZ', 'CS', '5Y', '6A',
    'AT', 'GG', 'G9', 'V7',
    'AV', 'GH', 'GA', 'YA', 'T5', 'AQ',
}

# UPS WORLDWIDE EXPRESS FREIGHT → deadline 14h
EXPRESS_FREIGHT_CODES = {
    'E1', 'E3', 'E4', 'E5', 'E9',
    'E2', 'E6', 'E7', 'E8', 'EA',
}

# UPS WORLDWIDE EXPRESS FREIGHT MIDDAY → deadline 14h
EXPRESS_FREIGHT_MIDDAY_CODES = {
    'EQ', 'ES', 'ET', 'EV', 'EZ',
    'ER', 'EW', 'EX', 'EY', 'F0',
}

# UPS EXPRESS SAVER (D9) → deadline 17h, traité comme standard (pas express)
EXPRESS_SAVER_CODES = {
    '04', '77', 'CH', 'CJ', 'D9', 'DA', '86', 'V6', '97', '93', 'DS', 'DV', '6G', '6H',
    'CE', 'CF', 'DD', 'DE',
}

# UPS EXPRESS 12:00 (Germany only) → deadline 14h
EXPRESS_1200_CODES = {
    'QH', 'Q4', 'QA', 'QC', 'QD', 'QE', 'Q5', 'Q8', 'Q6', 'Q7', 'Q2', 'Q3', 'Q0', 'Q1',
}

# Express = Express Plus + Express (deadline 14h) — Saver exclu (17h, pas express)
EXPRESS_SERVICE_CODES = (
    EXPRESS_PLUS_CODES | EXPRESS_CODES | EXPRESS_FREIGHT_CODES |
    EXPRESS_FREIGHT_MIDDAY_CODES | EXPRESS_1200_CODES
)

# Pour info : codes NON-express (standard, economy…)
# UPS STANDARD      : 68, FX, 79, GJ, CP, GK, CR, GL, DK, FY, DL, FZ, 91, 99, 95, DY, DZ, 6T, 6W
# UPS ACCESS POINT  : YZ, Z6, Z2, Z5, Z3, Z4, Z7
# UPS EXPEDITED     : 67, CL, CM, DG, DH, 88, 98, 94, DW, DX, 6N, 6P
# UPS ECONOMY DDP   : FC
# UPS ECONOMY DDU   : FD

# ─── Helpers temps ────────────────────────────────────────────────────────────

def _is_express_by_barcode(barcode: str) -> bool:
    """Détecte express via les 2 chiffres de service dans le barcode 1Z (pos 8-9)."""
    bc = str(barcode or '').strip()
    if len(bc) >= 10:
        return bc[8:10] in EXPRESS_SERVICE_CODES
    return False


def _service_code(barcode: str) -> str:
    """Extrait les 2 chiffres de service du barcode 1Z (positions 8-9)."""
    bc = str(barcode or '').strip()
    return bc[8:10] if len(bc) >= 10 else ''


def _commit_time(barcode: str, t_val: str, postal: str) -> Optional[dtime]:
    """
    Retourne l'heure limite selon les règles métier :
      CLOSED 1/2/3          → 17h00
      CP 2040               → 14h00 (port)
      Express Plus / Express→ 14h00
      Express Saver (D9…)  → 17h00 (traité comme standard)
      Standard              → 17h00
    """
    t_val  = str(t_val  or '').strip().upper()
    postal = str(postal or '').strip()
    sc     = _service_code(barcode)

    # CLOSED 1/2/3 → 17h
    if t_val in ('CLOSED 1', 'CLOSED 2', 'CLOSED 3'):
        return dtime(17, 0)

    # CP 2040 → 14h (zone port)
    if postal == '2040':
        return dtime(14, 0)

    # Express Plus / Express / Freight → 14h
    if sc in EXPRESS_SERVICE_CODES:
        return dtime(14, 0)

    # Express Saver → 17h (comme standard)
    # Standard / non-reconnu → 17h
    return dtime(17, 0)


def _parse_time(t_str) -> Optional[dtime]:
    """Parse 'HH:MM' ou 'HH:MM:SS' vers time object."""
    if not t_str:
        return None
    s = str(t_str).strip()
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def _is_express(barcode: str) -> bool:
    """Express = service code Express Plus ou Express dans barcode (Saver exclu)."""
    return _is_express_by_barcode(barcode)


def _apply_cluster_rule(rows: list, max_dist_m: float = 200.0, min_cluster: int = 3) -> list:
    """
    Règle des 3 express consécutifs :
    Si 3+ livraisons express se suivent et sont toutes à moins de max_dist_m
    les unes des autres (adresses côte à côte), le retard n'est pas sanctionné.
    """
    n = len(rows)
    excused = set()
    i = 0
    while i < n:
        r = rows[i]
        if not r['is_express'] or not r['gps_lat'] or not r['gps_lon']:
            i += 1
            continue
        # Construire cluster de express consécutifs géographiquement proches
        cluster = [i]
        j = i + 1
        while j < n:
            rj = rows[j]
            if not rj['is_express'] or not rj['gps_lat'] or not rj['gps_lon']:
                break
            prev = rows[cluster[-1]]
            dist = _haversine(float(prev['gps_lat']), float(prev['gps_lon']),
                              float(rj['gps_lat']),  float(rj['gps_lon']))
            if dist <= max_dist_m:
                cluster.append(j)
                j += 1
            else:
                break
        if len(cluster) >= min_cluster:
            for idx in cluster:
                excused.add(idx)
        i = j if len(cluster) > 1 else i + 1

    for idx in excused:
        rows[idx]['excused'] = True
        rows[idx]['is_late'] = False   # non sanctionné
    return rows


# ─── Géocodage Nominatim avec cache ──────────────────────────────────────────

def _load_geocache() -> dict:
    if os.path.exists(GEOCACHE):
        with open(GEOCACHE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def _save_geocache(cache: dict):
    os.makedirs(os.path.dirname(GEOCACHE), exist_ok=True)
    with open(GEOCACHE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def _geocode(street: str, city: str, postal: str, cache: dict) -> Optional[tuple]:
    """Retourne (lat, lon) ou None. Utilise le cache Nominatim."""
    key = f"{street}, {postal} {city}, Belgium"
    if key in cache:
        v = cache[key]
        return tuple(v) if v else None

    import urllib.request
    import urllib.parse
    query = urllib.parse.urlencode({'q': key, 'format': 'json', 'limit': '1'})
    url = f'https://nominatim.openstreetmap.org/search?{query}'
    req = urllib.request.Request(url, headers={'User-Agent': 'MD-Logistique/1.0 hajarbel325@gmail.com'})
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read())
        if data:
            lat, lon = float(data[0]['lat']), float(data[0]['lon'])
            cache[key] = (lat, lon)
            return (lat, lon)
    except Exception:
        pass
    cache[key] = None
    return None


def _haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Distance en mètres entre deux points GPS."""
    R = 6_371_000
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))


def _perimeter_color(dist_m: Optional[float]):
    """Retourne (bg, fg, label) selon la distance."""
    if dist_m is None:
        return BG_GREY, FG_GREY, '?'
    if dist_m <= 10:
        return BG_GREEN, FG_GREEN, f'{dist_m:.0f}m ✓'
    if dist_m <= 15:
        return BG_GREEN_LIGHT, FG_GREEN_LIGHT, f'{dist_m:.0f}m'
    if dist_m <= 30:
        return BG_YELLOW, FG_YELLOW, f'{dist_m:.0f}m'
    if dist_m <= 50:
        return BG_ORANGE, FG_ORANGE, f'{dist_m:.0f}m'
    return BG_RED, FG_RED, f'{dist_m:.0f}m !'


# ─── Liens Google Maps ────────────────────────────────────────────────────────

def _streetview_link(lat, lon) -> str:
    if lat and lon:
        return f'https://www.google.com/maps/@{lat},{lon},3a,75y,90t'
    return ''


def _maps_location(street, city, postal) -> str:
    """Lien Google Maps vers l'adresse de livraison (épingle)."""
    import urllib.parse
    addr = urllib.parse.quote(f'{street}, {postal} {city}, Belgium')
    return f'https://www.google.com/maps/search/{addr}'


def _check_timing_rule(rows: list, min_seconds: int = 80) -> list:
    """
    Règle timing : si deux stops CONSÉCUTIFS à des adresses DIFFÉRENTES
    sont complétés avec moins de min_seconds (1min20 = 80s) d'écart,
    marquer timing_warn=True (impossible de se déplacer aussi vite).
    """
    from datetime import datetime as dt

    def _to_sec(t_str):
        if not t_str:
            return None
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                t = dt.strptime(str(t_str).strip(), fmt)
                return t.hour * 3600 + t.minute * 60 + t.second
            except ValueError:
                continue
        return None

    for i, row in enumerate(rows):
        row['timing_warn'] = False
        if i == 0:
            continue
        prev = rows[i - 1]
        # Adresses différentes ?
        curr_addr = f"{row['street']} {row['postal']}"
        prev_addr = f"{prev['street']} {prev['postal']}"
        if curr_addr == prev_addr:
            continue
        t_curr = _to_sec(row['actual_time'])
        t_prev = _to_sec(prev['actual_time'])
        if t_curr is None or t_prev is None:
            continue
        diff = t_curr - t_prev
        if 0 <= diff < min_seconds:
            row['timing_warn'] = True
    return rows


def _add_consecutive_distances(rows: list, summary: dict) -> list:
    """
    Calcule la distance à vol d'oiseau entre l'adresse du stop N-1
    et l'adresse du stop N.  Met à jour dist_m et les compteurs summary.
    """
    summary.update({'dist_ok': 0, 'dist_warn': 0, 'dist_alarm': 0})
    for i, row in enumerate(rows):
        curr_coords = row.get('addr_coords')
        if i == 0 or curr_coords is None:
            row['dist_m'] = None
            continue
        prev_coords = rows[i - 1].get('addr_coords')
        if prev_coords is None:
            row['dist_m'] = None
            continue
        d = _haversine(prev_coords[0], prev_coords[1], curr_coords[0], curr_coords[1])
        row['dist_m'] = d
        if d <= 10:
            summary['dist_ok'] += 1
        elif d <= 30:
            summary['dist_warn'] += 1
        elif d > 50:
            summary['dist_alarm'] += 1
    return rows


def _add_consecutive_routes(rows: list) -> list:
    """
    Remplace le champ 'route' de chaque ligne par l'itinéraire
    entre le stop PRÉCÉDENT et le stop COURANT (adresse à adresse).
    Utilise le format Google Maps API (?api=1&origin=...&destination=...).
    → 1er stop : lien de recherche vers l'adresse courante uniquement.
    """
    import urllib.parse

    def _addr(r):
        return f"{r['street']}, {r['postal']} {r['city']}, Belgium"

    BASE = 'https://www.google.com/maps'

    for i, row in enumerate(rows):
        if not row['street']:
            row['route'] = ''
            continue
        curr_addr = _addr(row)
        if i == 0 or not rows[i - 1]['street']:
            # Premier stop : vue de l'adresse sur Maps
            q = urllib.parse.quote(curr_addr)
            row['route'] = f'{BASE}/search/?api=1&query={q}'
        else:
            prev_addr = _addr(rows[i - 1])
            o = urllib.parse.quote(prev_addr)
            d = urllib.parse.quote(curr_addr)
            row['route'] = f'{BASE}/dir/?api=1&origin={o}&destination={d}'
    return rows


# ─── Traitement principal ─────────────────────────────────────────────────────

def process_gps(
    file_path: str,
    output_path: Optional[str] = None,
    geocode: bool = True,
    progress_cb=None,
) -> dict:
    """
    Traite le fichier GPS et retourne:
    {
      'rows': [...],
      'summary': { 'total', 'late', 'express_late', 'dist_ok', 'dist_warn', 'dist_alarm' },
      'excel_bytes': bytes,
    }
    """
    wb_in = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws_in = wb_in.active
    raw_rows = list(ws_in.iter_rows(values_only=True))
    wb_in.close()

    if not raw_rows:
        return {'rows': [], 'summary': {}, 'excel_bytes': b''}

    headers = [str(h).strip() if h else '' for h in raw_rows[0]]

    # Index colonnes
    def col(name):
        return headers.index(name) if name in headers else None

    I_DATE   = col('Day Date')
    I_DRIVER = col('DIAD Orgnl Pkg Driver Name')
    I_BC     = col('Pkg Barcode Num')
    I_G      = col('Deliv Service Commit Time')
    I_H      = col('Street Name')
    I_I      = col('Street Num')
    I_J      = col('City Name')
    I_K      = col('Postal Code')
    I_L      = col('Stop Complete Time')
    I_M      = col('Stop Actual Sequence Num')
    I_T      = col('Deliv Info Status Reason Desc')
    I_Q      = col('Trigger 3 GPS Latitude')
    I_R      = col('Trigger 3 GPS Longitude')
    I_U      = col('Dispatch Loop Position Num')
    I_W      = col('DIAD Consignee Name')

    geocache = _load_geocache()  # toujours charger le cache existant
    geocache_dirty = False

    rows_out = []
    summary = {'total': 0, 'late': 0, 'express_late': 0,
                'dist_ok': 0, 'dist_warn': 0, 'dist_alarm': 0}

    data_rows = raw_rows[1:]
    n = len(data_rows)

    for idx, row in enumerate(data_rows):
        if progress_cb:
            progress_cb(idx, n)

        def v(i):
            return row[i] if i is not None and i < len(row) else None

        barcode  = str(v(I_BC) or '').strip()
        g_val    = v(I_G)
        t_val    = str(v(I_T) or '').strip()
        postal   = str(v(I_K) or '').strip()
        street   = str(v(I_H) or '').strip()
        street_n = str(v(I_I) or '').strip()
        city     = str(v(I_J) or '').strip()
        l_str    = str(v(I_L) or '').strip()
        gps_lat  = v(I_Q)
        gps_lon  = v(I_R)
        loop_val = str(v(I_U) or '').strip()

        # Exclure les colis du loop Q01
        if loop_val == 'Q01':
            continue

        full_street = f'{street} {street_n}'.strip()

        # Heure limite et heure réelle
        commit = _commit_time(barcode, t_val, postal)
        actual = _parse_time(l_str)

        is_exp  = _is_express(barcode)
        is_late = (commit is not None and actual is not None and actual > commit)

        summary['total'] += 1
        if is_late:
            summary['late'] += 1
        if is_late and is_exp:
            summary['express_late'] += 1

        # Distance GPS scan → adresse de livraison (géocodage Nominatim avec cache)
        dist_m = None
        addr_coords = None
        if full_street or city:
            key = f"{full_street}, {postal} {city}, Belgium"
            if key in geocache:
                raw = geocache[key]
                addr_coords = tuple(raw) if raw else None
            elif geocode:
                addr_coords = _geocode(full_street, city, postal, geocache)
                geocache_dirty = True
                time.sleep(1.1)  # Nominatim rate limit

            if addr_coords and gps_lat and gps_lon:
                dist_m = _haversine(float(gps_lat), float(gps_lon),
                                    addr_coords[0], addr_coords[1])

        if dist_m is not None:
            if dist_m <= 10:
                summary['dist_ok'] += 1
            elif dist_m <= 30:
                summary['dist_warn'] += 1
            elif dist_m > 50:
                summary['dist_alarm'] += 1

        # Liens Maps (route ajouté en post-traitement)
        sv_link = _streetview_link(gps_lat, gps_lon) if gps_lat else ''

        rows_out.append({
            'date':        v(I_DATE),
            'driver':      str(v(I_DRIVER) or '').strip(),
            'barcode':     barcode,
            'g_val':       g_val,
            'commit_time': commit.strftime('%H:%M') if commit else '',
            'actual_time': l_str,
            'excused':     False,
            'street':      full_street,
            'city':        city,
            'postal':      postal,
            'consignee':   str(v(I_W) or '').strip(),
            'status':      t_val,
            'seq':         str(v(I_M) or '').strip(),
            'loop':        str(v(I_U) or '').strip(),
            'gps_lat':     gps_lat,
            'gps_lon':     gps_lon,
            'is_express':  is_exp,
            'is_late':     is_late,
            'timing_warn': False,       # rempli par _check_timing_rule
            'dist_m':      dist_m,      # GPS scan → adresse géocodée
            'streetview':  sv_link,
            'route':       '',          # rempli par _add_consecutive_routes
        })

    if geocache_dirty:
        _save_geocache(geocache)

    # ── Règle 3 express consécutifs géographiquement proches ─────────────────
    rows_out = _apply_cluster_rule(rows_out)

    # ── Règle timing < 1min20 entre stops différents ─────────────────────────
    rows_out = _check_timing_rule(rows_out)

    # ── Itinéraire consécutif (stop N-1 → stop N) ────────────────────────────
    rows_out = _add_consecutive_routes(rows_out)

    # ── Génération Excel ──────────────────────────────────────────────────────
    excel_bytes = _build_excel(rows_out, summary)
    if output_path:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        with open(output_path, 'wb') as f:
            f.write(excel_bytes)

    return {'rows': rows_out, 'summary': summary, 'excel_bytes': excel_bytes}


def _fill(bg):
    return PatternFill(fill_type='solid', fgColor=bg)

def _font(fg, bold=False):
    return Font(color=fg, bold=bold)


def _build_excel(rows: list, summary: dict) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GPS Analysis'

    # ── Résumé ─────────────────────────────────────────────────────────────────
    ws.append(['GPS MD1 — Analyse livraisons'])
    ws.cell(1, 1).font = Font(bold=True, size=13)
    ws.append([])
    ws.append(['Total colis',       summary.get('total', 0)])
    ws.append(['Retards (total)',    summary.get('late', 0)])
    ws.append(['Retards express',    summary.get('express_late', 0)])
    ws.append(['GPS OK (≤10m)',      summary.get('dist_ok', 0)])
    ws.append(['GPS warning (≤30m)', summary.get('dist_warn', 0)])
    ws.append(['GPS alarme (>50m)',  summary.get('dist_alarm', 0)])
    ws.append([])

    # ── Header ──────────────────────────────────────────────────────────────────
    col_headers = [
        'Date', 'Chauffeur', 'Barcode', 'Commit\n(contractuel)',
        'Réel\n(col L)', 'RETARD?', 'Express?',
        'Séquence', 'Loop', 'Statut (T)',
        'Rue', 'Ville', 'CP', 'Destinataire',
        'GPS Lat', 'GPS Lon',
        'Distance\nGPS→Adresse', 'Périmètre',
        'Itinéraire GPS→Adresse',
    ]
    ws.append(col_headers)
    hdr_row = ws.max_row
    for c in range(1, len(col_headers)+1):
        cell = ws.cell(hdr_row, c)
        cell.font      = Font(bold=True, color='FFFFFF', size=10)
        cell.fill      = PatternFill(fill_type='solid', fgColor='2F4F4F')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # ── Données ─────────────────────────────────────────────────────────────────
    for r in rows:
        late_str    = 'OUI' if r['is_late'] else ('EXCUSÉ' if r.get('excused') else '')
        express_str = 'EXPRESS' if r['is_express'] else 'STD'

        dist_bg, dist_fg, dist_label = _perimeter_color(r['dist_m'])

        date_val = r['date']
        if hasattr(date_val, 'strftime'):
            date_val = date_val.strftime('%d/%m/%Y')

        row_data = [
            date_val,
            r['driver'],
            r['barcode'],
            r['commit_time'],
            r['actual_time'],
            late_str,
            express_str,
            r['seq'],
            r['loop'],
            r['status'],
            r['street'],
            r['city'],
            r['postal'],
            r['consignee'],
            r['gps_lat'],
            r['gps_lon'],
            f"{r['dist_m']:.1f}" if r['dist_m'] is not None else '',
            dist_label,
            r['route'],
        ]
        ws.append(row_data)
        data_row = ws.max_row

        # Couleur de ligne : fond ORANGE + texte ROUGE GRAS si express EN RETARD
        if r['is_late'] and r['is_express']:
            for c in range(1, len(col_headers)+1):
                ws.cell(data_row, c).fill = _fill(BG_LATE_EXPRESS)
                ws.cell(data_row, c).font = Font(color=FG_LATE_EXPRESS, bold=True, size=10)

        # Couleur cellule distance (colonne 18)
        dist_cell = ws.cell(data_row, 18)
        dist_cell.fill = _fill(dist_bg)
        dist_cell.font = _font(dist_fg, bold=True)
        dist_cell.alignment = Alignment(horizontal='center')

        # Lien itinéraire cliquable (colonne 19)
        if r['route']:
            ws.cell(data_row, 19).hyperlink = r['route']
            ws.cell(data_row, 19).value     = 'Itinéraire'
            ws.cell(data_row, 19).font      = Font(color='0563C1', underline='single')

    # ── Largeurs ─────────────────────────────────────────────────────────────
    widths = [12, 14, 22, 10, 10, 8, 8, 8, 6, 14, 28, 14, 7, 22, 12, 12, 14, 12, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A11'

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CLI ─────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'samples/hakim/GPS MD1 M-D Logistique.xlsx'
    geo  = '--no-geo' not in sys.argv
    out  = 'outputs/GPS_rapport.xlsx'

    print(f'Traitement: {path}')
    print(f'Géocodage: {"OUI (Nominatim)" if geo else "NON"}')

    def prog(i, n):
        if i % 50 == 0:
            print(f'  {i}/{n} lignes...', end='\r')

    result = process_gps(path, output_path=out, geocode=geo, progress_cb=prog)
    s = result['summary']
    print(f'\n{"="*50}')
    print(f'  Total colis      : {s["total"]}')
    print(f'  Retards          : {s["late"]}')
    print(f'  Retards express  : {s["express_late"]}')
    print(f'  GPS OK (≤10m)    : {s["dist_ok"]}')
    print(f'  GPS warning      : {s["dist_warn"]}')
    print(f'  GPS alarme (>50m): {s["dist_alarm"]}')
    print(f'{"="*50}')
    print(f'  Rapport → {out}')
