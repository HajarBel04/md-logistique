#!/usr/bin/env python3
"""
MD-Logistique — Module B : Automatisation des fiches de paie CP140
Traitement Webfleet → CP140 + rapport récapitulatif Excel
"""

import os
import re
import json
import math
import zipfile
import io
import shutil
from datetime import datetime, date
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ─── Constantes ───────────────────────────────────────────────────────────────

JOURS_FERIES_MAI_2026 = {1, 14, 25}   # 1er mai, Ascension (14), Pentecôte (25)

# Colonnes CP140 (index openpyxl = lettre → numéro)
COL = {
    'date':       1,   # A
    'horaire':    2,   # B
    'code_abs':   3,   # C
    'travail':    4,   # D
    'dispo':      5,   # E
    'train_lt4':  6,   # F  heures train/bateau < 4h
    'service':    7,   # G
    'train_ge4':  8,   # H  heures train/bateau >= 4h
    'sejour_a':   9,   # I
    'sejour_b':   10,  # J
    'sejour_c':   11,  # K  blocage étranger
    'nuit':       12,  # L
    'equipe':     13,  # M
}

# ─── Patch openpyxl : corrige les <fill/> vides dans styles.xml ──────────────

def _fix_xlsx_styles(path_or_bytes):
    """
    Lit le fichier xlsx, corrige les éléments <fill/> vides qui font
    crasher openpyxl, retourne un BytesIO propre prêt à charger.
    """
    buf = io.BytesIO()
    src = open(path_or_bytes, 'rb') if isinstance(path_or_bytes, str) else path_or_bytes

    with zipfile.ZipFile(src, 'r') as zin, zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == 'xl/styles.xml':
                xml = data.decode('utf-8')
                # Remplace <fill/> par un fill valide
                xml = re.sub(r'<fill/>', '<fill><patternFill patternType="none"/></fill>', xml)
                data = xml.encode('utf-8')
            zout.writestr(item, data)

    if isinstance(path_or_bytes, str):
        src.close()

    buf.seek(0)
    return buf


def load_wb(path, data_only=False):
    """Charge un workbook en patchant les styles si nécessaire."""
    try:
        return openpyxl.load_workbook(path, data_only=data_only)
    except TypeError:
        buf = _fix_xlsx_styles(path)
        return openpyxl.load_workbook(buf, data_only=data_only)


# ─── Parsing durée ─────────────────────────────────────────────────────────────

def parse_duration_min(text) -> float:
    """'2 h 15 min' → 135.0  |  '3 min' → 3.0  |  '1 h' → 60.0"""
    if not text:
        return 0.0
    s = str(text).strip()
    h = re.search(r'(\d+)\s*h', s)
    m = re.search(r'(\d+)\s*min', s)
    return float(h.group(1)) * 60 if h else 0.0 + (float(m.group(1)) if m else 0.0)


def h(minutes: float) -> float:
    """Minutes → heures décimales (2 décimales)."""
    return round(minutes / 60, 2)


# ─── Lecture fichier Webfleet ─────────────────────────────────────────────────

def _parse_dt(s: str):
    """'01/05/26 22:49' ou '01/05/2026 22:49' → datetime"""
    s = str(s).strip()
    for fmt in ('%d/%m/%y %H:%M', '%d/%m/%Y %H:%M'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def read_webfleet(path: str) -> list:
    """
    Lit un fichier Webfleet .xlsx via XML direct (contourne le bug openpyxl fills).
    Retourne liste de dicts : activite, debut, fin, duree_min, pos_debut, pos_fin.
    """
    NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'

    with zipfile.ZipFile(path) as z:
        # Shared strings
        ss = []
        if 'xl/sharedStrings.xml' in z.namelist():
            from xml.etree.ElementTree import parse as xparse
            root = xparse(z.open('xl/sharedStrings.xml')).getroot()
            for si in root.findall(f'{{{NS}}}si'):
                t = si.find(f'{{{NS}}}t')
                if t is not None:
                    ss.append(t.text or '')
                else:
                    ss.append(''.join(
                        r.find(f'{{{NS}}}t').text or ''
                        for r in si.findall(f'{{{NS}}}r')
                        if r.find(f'{{{NS}}}t') is not None
                    ))

        # Sheet data
        from xml.etree.ElementTree import parse as xparse
        root = xparse(z.open('xl/worksheets/sheet1.xml')).getroot()
        raw_rows = []
        for row_el in root.findall(f'.//{{{NS}}}row'):
            row = {}
            for c in row_el.findall(f'{{{NS}}}c'):
                col = ''.join(ch for ch in c.get('r') if ch.isalpha())
                v = c.find(f'{{{NS}}}v')
                if v is not None and v.text:
                    if c.get('t') == 's':
                        row[col] = ss[int(v.text)]
                    else:
                        try:
                            row[col] = float(v.text) if '.' in v.text else int(v.text)
                        except ValueError:
                            row[col] = v.text
            raw_rows.append(row)

    activities = []
    for row in raw_rows[1:]:          # skip header row
        act = str(row.get('A', '')).strip()
        if not act or act == 'Activité':
            continue
        activities.append({
            'activite':  act,
            'debut':     _parse_dt(row.get('B', '')),
            'fin':       _parse_dt(row.get('C', '')),
            'duree_min': parse_duration_min(row.get('D')),
            'pos_debut': str(row.get('F', '')).strip(),
            'pos_fin':   str(row.get('G', '')).strip(),
        })

    return activities


# ─── Nom conducteur depuis nom de fichier ─────────────────────────────────────

def driver_name_from_file(path: str) -> str:
    """
    'Temps_de_travail_mohamed_sidibe_sidibe_2026-06-14_23-47.xlsx'
    → 'Mohamed Sidibe Sidibe'
    """
    name = os.path.basename(path)
    name = re.sub(r'\.xlsx$', '', name, flags=re.IGNORECASE)
    name = re.sub(r'^Temps_de_travail_', '', name, flags=re.IGNORECASE)
    name = re.sub(r'_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}$', '', name)
    return ' '.join(p.capitalize() for p in name.split('_') if p)


# ─── Règles métier ────────────────────────────────────────────────────────────

def classify_position(position: str, config: dict) -> str:
    """
    'travail' si adresse client connue
    'exclu'   si périmètre domicile (sans GPS on ne peut pas, retourne 'repos')
    'repos'   sinon
    """
    if not position or not config:
        return 'repos'
    pl = position.lower()
    for addr in config.get('adresses_clients', []):
        if addr.lower() in pl or pl in addr.lower():
            return 'travail'
    return 'repos'


def apply_pause_rule(activities: list) -> tuple:
    """
    Règle Alex : après 4h30 conduite continue → déduire 45 min de pause.
    Retourne (minutes_déduites, liste_anomalies).
    """
    sorted_acts = sorted(
        [a for a in activities if a['debut']],
        key=lambda x: x['debut']
    )

    deducted = 0.0
    anomalies = []
    conduite_seq = 0.0
    need_pause = False
    i = 0

    while i < len(sorted_acts):
        act = sorted_acts[i]
        if act['activite'] == 'Conduite':
            conduite_seq += act['duree_min']
            if conduite_seq >= 270:           # 4h30 = 270 min
                need_pause = True
        else:
            if need_pause:
                if act['activite'] in ('Repos', 'Disponibilité') and act['duree_min'] >= 45:
                    deducted += 45
                    need_pause = False
                    conduite_seq = 0.0
                else:
                    # Pause insuffisante — on continue à chercher la prochaine
                    pass
            if act['activite'] not in ('Conduite',):
                conduite_seq = 0.0
        i += 1

    if need_pause:
        anomalies.append("Pause 45min manquante après ≥4h30 conduite")

    return deducted, anomalies


def calculate_day(day_date: date, activities: list, config: dict) -> dict:
    """
    Calcule toutes les colonnes CP140 pour un jour donné.
    """
    day_num = day_date.day
    is_weekend = day_date.weekday() >= 5
    is_ferie   = day_num in JOURS_FERIES_MAI_2026

    result = {
        'date':        day_date,
        'day_num':     day_num,
        'is_weekend':  is_weekend,
        'is_ferie':    is_ferie,
        'travail_min': 0.0,
        'dispo_min':   0.0,
        'exclu_min':   0.0,
        'code_abs':    'F' if is_ferie else '',
        'anomalies':   [],
    }

    if not activities:
        return result

    pause_ded, pause_anom = apply_pause_rule(activities)
    result['anomalies'].extend(pause_anom)

    travail = dispo = exclu = 0.0

    for act in activities:
        t   = act['activite']
        dur = act['duree_min']

        if t == 'Conduite':
            travail += dur

        elif t == 'Travail':
            cls = classify_position(act['pos_debut'], config)
            if cls == 'exclu':
                exclu += dur
            else:
                travail += dur           # client ou inconnu → travail

        elif t == 'Repos':
            cls = classify_position(act['pos_debut'], config)
            if cls == 'exclu':
                exclu += dur
            else:
                dispo += dur

        elif t == 'Disponibilité':
            dispo += dur

    # Déduction pause obligatoire
    dispo = max(0.0, dispo - pause_ded)

    result['travail_min'] = travail
    result['dispo_min']   = dispo
    result['exclu_min']   = exclu
    return result


# ─── CP140 : trouver le bloc conducteur ──────────────────────────────────────

def _name_tokens(s: str) -> set:
    return set(re.sub(r"[^a-z0-9 ]", '', s.lower()).split())


def find_driver_base_row(ws, driver_name: str) -> int:
    """
    Cherche la ligne 'Travailleur:' pour le conducteur donné.
    Retourne le numéro de ligne ou -1.
    """
    target = _name_tokens(driver_name)

    for row in ws.iter_rows(min_col=1, max_col=4):
        if row[0].value == 'Travailleur:':
            name_val = row[3].value           # col D
            if name_val:
                cp_tokens = _name_tokens(str(name_val))
                overlap = target & cp_tokens
                # Accepte si au moins 2 tokens communs (ou tous si < 2 tokens)
                if len(overlap) >= min(2, len(target)):
                    return row[0].row
    return -1


def build_day_row_map(ws, base_row: int) -> dict:
    """
    Scanne le bloc conducteur et construit {jour: row_excel}.
    """
    day_map = {}
    for offset in range(5, 52):
        row_num = base_row + offset
        cell = ws.cell(row=row_num, column=COL['date'])
        if cell.value:
            m = re.match(r'^(\d{2})\s+\w+', str(cell.value))
            if m:
                day_map[int(m.group(1))] = row_num
    return day_map


# ─── Écriture dans le CP140 ───────────────────────────────────────────────────

def write_driver_block(ws, base_row: int, day_results: list):
    """Remplit les cellules jour par jour dans le bloc du conducteur."""
    day_map = build_day_row_map(ws, base_row)

    for dr in day_results:
        row = day_map.get(dr['day_num'])
        if row is None or dr['is_weekend']:
            continue

        travail_h = h(dr['travail_min'])
        dispo_h   = h(dr['dispo_min'])
        service_h = round(travail_h + dispo_h, 2)

        if dr['code_abs']:
            ws.cell(row=row, column=COL['code_abs']).value = dr['code_abs']

        ws.cell(row=row, column=COL['travail']).value = travail_h
        ws.cell(row=row, column=COL['dispo']).value   = dispo_h
        ws.cell(row=row, column=COL['service']).value = service_h
        # F / H (heures train/bateau) : laissées vides sans config ferry


# ─── Rapport récapitulatif Excel ──────────────────────────────────────────────

def generate_recap(driver_name: str, day_results: list, output_path: str) -> dict:
    """Génère le fichier récapitulatif par conducteur."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = driver_name[:31]

    # Palettes couleurs
    C_HEADER = PatternFill('solid', fgColor='1E3A5F')
    C_GREEN  = PatternFill('solid', fgColor='C6EFCE')
    C_ORANGE = PatternFill('solid', fgColor='FFEB9C')
    C_YELLOW = PatternFill('solid', fgColor='FFFACD')
    C_RED    = PatternFill('solid', fgColor='FFC7CE')
    C_TOTAL  = PatternFill('solid', fgColor='BDD7EE')

    HEADERS = [
        'Date', 'Jour', 'Code Abs', 'Travail (h)', 'Dispo (h)',
        'Service (h)', 'Exclu (h)', 'Anomalie'
    ]
    JOURS_FR = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di']

    # En-tête titre
    ws.cell(row=1, column=1, value=f"Récapitulatif — {driver_name} — Mai 2026").font = Font(bold=True, size=13)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # En-têtes colonnes
    for c, hdr in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=hdr)
        cell.fill = C_HEADER
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center')

    totals = {'travail': 0.0, 'dispo': 0.0, 'jours': 0, 'anomalies': 0}

    for i, dr in enumerate(day_results, 3):
        jour = JOURS_FR[dr['date'].weekday()]
        tv   = h(dr['travail_min'])
        dv   = h(dr['dispo_min'])
        sv   = round(tv + dv, 2)
        anom = '; '.join(dr['anomalies']) if dr['anomalies'] else ''

        row_data = [
            dr['date'].strftime('%d/%m/%Y'), jour, dr['code_abs'],
            tv, dv, sv, h(dr['exclu_min']), anom
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=i, column=c, value=val)

        # Couleur de ligne
        if dr['is_ferie']:
            fill = C_RED
        elif dr['is_weekend']:
            fill = C_YELLOW
        elif anom:
            fill = C_ORANGE
        else:
            fill = C_GREEN

        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=i, column=c).fill = fill

        if not dr['is_weekend'] and not dr['is_ferie']:
            totals['travail'] += dr['travail_min']
            totals['dispo']   += dr['dispo_min']
            if tv > 0 or dv > 0:
                totals['jours'] += 1
            if anom:
                totals['anomalies'] += 1

    # Ligne total
    tr = len(day_results) + 3
    totals_row = ['TOTAL', '', '', h(totals['travail']), h(totals['dispo']),
                  h(totals['travail'] + totals['dispo']), '', f"{totals['anomalies']} anomalie(s)"]
    for c, val in enumerate(totals_row, 1):
        cell = ws.cell(row=tr, column=c, value=val)
        cell.font = Font(bold=True)
        cell.fill = C_TOTAL

    # Largeur colonnes automatique
    for col_idx, col_letter in enumerate(
        ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], 1
    ):
        col_cells = [ws.cell(row=r, column=col_idx) for r in range(1, tr + 1)]
        width = max((len(str(c.value or '')) for c in col_cells), default=8)
        ws.column_dimensions[col_letter].width = width + 4

    wb.save(output_path)

    return {
        'jours_travailles': totals['jours'],
        'total_travail_h':  h(totals['travail']),
        'total_service_h':  h(totals['travail'] + totals['dispo']),
        'nb_anomalies':     totals['anomalies'],
    }


# ─── Traitement complet d'un conducteur ──────────────────────────────────────

def process_driver(
    webfleet_path: str,
    cp140_path: str,         # chemin du fichier CP140 à modifier (copie de travail)
    config: dict,
    output_recap: str,
) -> dict:
    """
    Traite un conducteur : lit Webfleet, calcule les jours, écrit CP140 + récap.
    """
    name = driver_name_from_file(webfleet_path)
    print(f"\n{'─'*60}")
    print(f"  Conducteur : {name}")
    print(f"  Fichier    : {os.path.basename(webfleet_path)}")
    print(f"{'─'*60}")

    # 1. Lecture activités Webfleet
    activities = read_webfleet(webfleet_path)
    print(f"  {len(activities)} lignes d'activités lues")

    # 2. Groupement par jour
    days_map = defaultdict(list)
    for act in activities:
        if act['debut']:
            days_map[act['debut'].date()].append(act)

    # 3. Calcul jour par jour (mai 2026 = 31 jours)
    day_results = []
    for d in range(1, 32):
        dd = date(2026, 5, d)
        acts = days_map.get(dd, [])
        dr = calculate_day(dd, acts, config)
        day_results.append(dr)

        if dr['travail_min'] > 0 or dr['dispo_min'] > 0:
            jour = ['Lu', 'Ma', 'Me', 'Je', 'Ve', 'Sa', 'Di'][dd.weekday()]
            tv = h(dr['travail_min'])
            dv = h(dr['dispo_min'])
            sv = round(tv + dv, 2)
            flag = '  ⚠ ' + '; '.join(dr['anomalies']) if dr['anomalies'] else ''
            print(f"  {dd.strftime('%d/%m')} {jour}: travail={tv}h  dispo={dv}h  service={sv}h{flag}")

    # 4. Écriture dans le CP140
    print(f"\n  Recherche dans CP140…")
    wb = load_wb(cp140_path)
    ws = wb['Etat de prestation']
    base_row = find_driver_base_row(ws, name)

    driver_found = base_row != -1
    if not driver_found:
        print(f"  ⚠  '{name}' non trouvé dans le CP140 (vérifier la correspondance des noms)")
    else:
        print(f"  ✓  Bloc trouvé à la ligne {base_row}")
        write_driver_block(ws, base_row, day_results)
        wb.save(cp140_path)
        print(f"  ✓  CP140 mis à jour")

    # 5. Rapport récapitulatif
    totals = generate_recap(name, day_results, output_recap)
    print(f"  ✓  Récap sauvegardé : {output_recap}")

    # 6. Collecte anomalies toutes dates
    all_anomalies = [
        f"{dr['date'].strftime('%d/%m')} : {a}"
        for dr in day_results
        for a in dr['anomalies']
    ]

    return {
        'nom':                name,
        'driver_found':       driver_found,
        'base_row':           base_row,
        'jours_travailles':   totals['jours_travailles'],
        'total_heures_travail': totals['total_travail_h'],
        'total_heures_service': totals['total_service_h'],
        'anomalies':          all_anomalies,
        'status':             'ok' if driver_found else 'driver_not_found',
    }


# ─── Point d'entrée CLI ───────────────────────────────────────────────────────

def main():
    import sys

    BASE    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    SAMPLES = os.path.join(BASE, 'samples')
    OUTPUTS = os.path.join(BASE, 'outputs')
    os.makedirs(OUTPUTS, exist_ok=True)

    # Config
    config_path = os.path.join(BASE, 'documents', 'config.json')
    config = {}
    if os.path.exists(config_path):
        with open(config_path) as f:
            config = json.load(f)

    # Template CP140 → copie de travail dans outputs/
    template = os.path.join(SAMPLES, 'alex', 'PC140_991163 01-05-2026 - 31-05-2026.xlsx')
    cp140_out = os.path.join(OUTPUTS, 'CP140_rempli_mai2026.xlsx')
    shutil.copy2(template, cp140_out)
    print(f"Copie template → {cp140_out}")

    # Fichiers Webfleet
    if len(sys.argv) > 1:
        wf_files = sys.argv[1:]
    else:
        wf_files = [
            os.path.join(SAMPLES, 'alex', 'rapport_juin',
                         'Temps_de_travail_mohamed_sidibe_sidibe_2026-06-14_23-47.xlsx')
        ]

    results = []
    for wf in wf_files:
        slug = re.sub(r'^Temps_de_travail_', '', os.path.splitext(os.path.basename(wf))[0])
        slug = re.sub(r'_\d{4}-\d{2}-\d{2}_\d{2}-\d{2}$', '', slug)
        recap_out = os.path.join(OUTPUTS, f'recap_{slug}.xlsx')

        result = process_driver(wf, cp140_out, config, recap_out)
        results.append(result)

    # Résumé final
    print(f"\n{'='*60}")
    print("  RÉSUMÉ FINAL")
    print(f"{'='*60}")
    for r in results:
        ok = '✓' if r['driver_found'] else '✗'
        print(f"\n  {ok} {r['nom']}")
        print(f"     Jours travaillés    : {r['jours_travailles']}")
        print(f"     Total heures travail : {r['total_heures_travail']}h")
        print(f"     Total heures service : {r['total_heures_service']}h")
        if r['anomalies']:
            print(f"     Anomalies ({len(r['anomalies'])}):")
            for a in r['anomalies']:
                print(f"       ⚠  {a}")

    print(f"\n  CP140 final : {cp140_out}")
    print(f"{'='*60}\n")
    return results


if __name__ == '__main__':
    main()
